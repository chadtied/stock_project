from seleniumbase import Driver
from selenium.webdriver.common.by import By
import openpyxl


#設定excel
workbook= openpyxl.Workbook()
workpage= workbook.create_sheet("List1",0)
workpage.cell(1,1).value= "指標名稱"
workpage.cell(1,2).value= "開盤"
workpage.cell(1,3).value= "最高"
workpage.cell(1,4).value= "最低"
workpage.cell(1,5).value= "收盤"
workpage.cell(1,6).value= "成交量(億)"

#建立dict存取資訊
imfor= dict()

#建立變數
driver= Driver(uc= True, incognito= False, headless= True)
tiles_set= {"開盤","最高","最低","昨收","成交量(億)"}
urls= [
    'https://www.wantgoo.com/index/0000',
    'https://www.wantgoo.com/global/dji'
]
row, col= 2, 1

for url in urls:
    driver.get(url)
    workpage.cell(row, 1).value= driver.find_element(By.CSS_SELECTOR,"#investrue-info-1 > h3").text
    figures= driver.find_elements(By.CSS_SELECTOR,"body > div.page-wrap > main > div > div.quotes-wrap > div.quotes-info > div.lasty-detail > ul > li")
    col= 2
    for figure in figures:
        key= figure.find_element(By.CSS_SELECTOR,"i").text
        val= figure.find_element(By.CSS_SELECTOR,"span").text  
        if key in tiles_set: 
            workpage.cell(row, col).value= val
            col+= 1
    row+= 1


workbook.save("./new.xlsx")